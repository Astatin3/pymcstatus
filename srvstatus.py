import io
import math
import base64
import sys
from time import sleep
from threading import Thread
from PIL import Image
from mcstatus import JavaServer

from openpyxl import load_workbook
from openpyxl import Workbook


imgsize = (16,16)
sleeptime = 0.1
nullimg = b"iVBORw0KGgoAAAANSUhEUgAAABAAAAAQCAYAAAAf8/9hAAAAaklEQVQ4T62TWw7AIAgEl/sf2qoRY4ywa1tD/JvhoRjGKUAN/RhQY1y3sKdpEltht2Z17Mmm4A3cW3AjE+yVTk4RnNqUBdGMJEE2YCpgr5MKGNymHwoUOBSo8FGw/j72J1LBDfzPMn1d5wfNZUf5qKNxAQAAAABJRU5ErkJggg=="
threadcount = 1

outputfilename = None

def appenddata(ip, latency, pver, ver, cplayers, mplayers, motd, isicon):
    book = None
    sheet = None
    try:
        book = load_workbook(filename=outputfilename)
        sheet = book.active
    except:
        book = Workbook()
        sheet = book.active
        sheet.append(("ip", "latency", "pver", "ver", "cplayers", "mplayers", "motd", "isicon"))


    sheet.append((ip, latency, pver, ver, cplayers, mplayers, motd, isicon))

    book.save(outputfilename)


def img2ascii(icon):

    img = Image.open(io.BytesIO(base64.b64decode(icon)))
    img = img.resize(imgsize)
    imageSizeW, imageSizeH = img.size

    asciiimg = []
    for i in range(0, imageSizeW):
        curline = ""
        for j in range(0, imageSizeH):
            pixVal = img.getpixel((j, i))
            
            curline += (f"\033[38;2;{pixVal[0]};{pixVal[1]};{pixVal[2]}m██")

        asciiimg.append(curline+"\033[38;2;255;255;255m > ")
    return asciiimg

def scanip(ip):
    global threadcount
    threadcount += 1
    try:
        server = JavaServer.lookup(ip)
        status = server.status()
    except:
        threadcount -= 1
        return False
        
    isicon = None

    if status.icon == None:
        b64img = nullimg
        isicon = False
    else:
        b64img = status.icon[22:]
        isicon = True
    
    try: # In case an icon is invalid
        ascii = img2ascii(b64img)
    except:
        ascii = img2ascii(nullimg)
        isicon = False

    ver = status.version
    players = status.players

    motdtxt = status.motd.to_plain()

    if outputfilename != None:
        appenddata(ip, status.latency, ver.protocol, ver.name, players.online, players.max, motdtxt, isicon)

    motd = ["", ""]
    if "\n" in motdtxt:
        split = motdtxt.split("\n")
        motd[0] = split[0]
        motd[1] = split[1]
    else:
        motd[0] = motdtxt

    output = ""
    for ln in range(0,imgsize[1],1):
        line = ascii[ln]
        match ln:
            case 0:
                print("\r\033[K", end="")
                line += f"IP: {ip}"
            case 1:
                line += f"Latency: {math.ceil(status.latency)}ms"
            case 2:
                line += f"P: {ver.protocol}, V: {ver.name}"
            case 3:
                line += f"{players.online}/{players.max} Players"
            case 5:
                line += f"{motd[0]}"
            case 6:
                line += f"{motd[1]}"

        output += f"{line}\n"
    #with open("output.txt", "a") as file:
    #    file.write(f"{output}\n")
    print(f"{output}\n") # If everything isn't printed at the same time, the output text could become mixed.
    threadcount -= 1
    return True

threads = []

if sys.argv[1] == "-p":
    if not scanip(sys.argv[2]):
        print("Server not online")
elif sys.argv[1] == "-f":
    if sys.argv[3] == "-o":
        outputfilename = sys.argv[4]

    with open(sys.argv[2]) as file:
        length = sum(1 for line in open(sys.argv[2]))
        for index, ip in enumerate(file):

            ip = ip.rstrip()

            ipgap = ' '*(15-len(ip))
            per = round(100*(index/length), 2)
            pergap = ' '*(6-len(str(per)))
            thrgap = ' '*(3-len(str(threadcount)))

            print(f"Scanning IP: {ip}{ipgap} {pergap}{per}% done {thrgap} {threadcount} pending", end='\r')
            thread = Thread(target = scanip, args = (ip, ))
            thread.start()
            threads.append(thread)
            sleep(sleeptime)

    scanip(sys.argv[1])

    for thread in threads:
        thread.join()
else:
    print("Usage:\n"+\
          "srvstatus.py -p IP\n"+\
          "srvstatus.py -f inlist.txt\n"+\
          "srvstatus.py -f inlist.txt -o outfile.xlsx")
